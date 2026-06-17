"""Multi-format file/document metadata extractor.

Supports: images (EXIF+GPS+geocode), audio (ID3/Vorbis), video (hachoir),
          PDF, DOCX, XLSX, generic.
All analysis is local where possible; GPS reverse-geocoding uses Nominatim OSM (free, no key).
"""
import datetime
import hashlib
import mimetypes
import os
import stat
from pathlib import Path
from typing import Any

try:
    from PIL import Image, ExifTags
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False


def _human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


def _file_hashes(path: Path) -> dict:
    md5 = hashlib.md5()
    sha256 = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                md5.update(chunk)
                sha256.update(chunk)
        return {"md5": md5.hexdigest(), "sha256": sha256.hexdigest()}
    except Exception:
        return {"md5": "", "sha256": ""}


def _file_base(path: Path) -> dict:
    s = path.stat()
    mime, _ = mimetypes.guess_type(str(path))
    hashes = _file_hashes(path)
    # st_birthtime is available on macOS; Linux falls back to st_ctime (inode change)
    created_ts = getattr(s, "st_birthtime", s.st_ctime)
    return {
        "file_name": path.name,
        "file_size_bytes": s.st_size,
        "file_size_human": _human_size(s.st_size),
        "extension": path.suffix.lower(),
        "mime_type": mime or "application/octet-stream",
        "md5": hashes["md5"],
        "sha256": hashes["sha256"],
        "fs_modified": datetime.datetime.fromtimestamp(s.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "fs_accessed": datetime.datetime.fromtimestamp(s.st_atime).strftime("%Y-%m-%d %H:%M:%S"),
        "fs_created": datetime.datetime.fromtimestamp(created_ts).strftime("%Y-%m-%d %H:%M:%S"),
    }


# Simple in-process cache keyed on (lat, lon) rounded to 4 decimal places (~11 m).
# Prevents duplicate Nominatim calls for the same location and respects the
# Nominatim usage policy of max 1 request/second per IP.
_geocode_cache: dict = {}


def _reverse_geocode(lat: float, lon: float) -> str | None:
    """Look up a human-readable address from GPS coords via Nominatim (free, no key)."""
    key = (round(lat, 4), round(lon, 4))
    if key in _geocode_cache:
        return _geocode_cache[key]
    try:
        import httpx
        r = httpx.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={"lat": lat, "lon": lon, "format": "json"},
            headers={"User-Agent": "OSINT-Tracker/1.0 (ethical research)"},
            timeout=5,
        )
        if r.status_code == 200:
            data = r.json()
            result = data.get("display_name") or None
            _geocode_cache[key] = result
            return result
    except Exception:
        pass
    return None


def _image(path: Path) -> dict:
    result = {"file_type": "image", **_file_base(path)}
    if not PILLOW_AVAILABLE:
        result["error"] = "Pillow not installed"
        return result
    try:
        with Image.open(path) as img:
            result["format"] = img.format or path.suffix.lstrip(".")
            result["mode"] = img.mode
            result["dimensions"] = f"{img.width}x{img.height}"
            result["width"] = img.width
            result["height"] = img.height

            exif: dict[str, Any] = {}
            gps_lat = gps_lon = None

            try:
                raw = img.getexif()
                if raw:
                    GPS_IFD_TAG = 0x8825

                    # --- GPS: use get_ifd() — Pillow 10+ returns GPSInfo as an
                    # integer offset when iterating getexif(), not a dict.
                    try:
                        from PIL.ExifTags import GPSTAGS
                        gps_ifd = raw.get_ifd(GPS_IFD_TAG)
                        if gps_ifd:
                            gps = {GPSTAGS.get(k, k): v for k, v in gps_ifd.items()}
                            lat_raw = gps_ifd.get(2)   # GPSLatitude
                            lon_raw = gps_ifd.get(4)   # GPSLongitude
                            lat_ref = str(gps_ifd.get(1, "N"))
                            lon_ref = str(gps_ifd.get(3, "E"))
                            if lat_raw and lon_raw:
                                def _dms_to_decimal(dms) -> float:
                                    """Convert DMS tuple/list to decimal degrees.
                                    Handles IFDRational, float, int, and tuple elements."""
                                    parts = list(dms)
                                    if len(parts) < 3:
                                        raise ValueError("DMS needs 3 elements")
                                    return float(parts[0]) + float(parts[1]) / 60 + float(parts[2]) / 3600

                                ld = _dms_to_decimal(lat_raw)
                                lo = _dms_to_decimal(lon_raw)
                                if lat_ref.upper() == "S":
                                    ld = -ld
                                if lon_ref.upper() == "W":
                                    lo = -lo
                                gps_lat, gps_lon = ld, lo
                                exif["GPS_Coordinates"] = f"{ld:.6f}, {lo:.6f}"
                                exif["GPS_Latitude"] = f"{ld:.6f} ({lat_ref})"
                                exif["GPS_Longitude"] = f"{lo:.6f} ({lon_ref})"
                                alt = gps.get("GPSAltitude")
                                if alt is not None:
                                    exif["GPS_Altitude"] = str(alt)
                    except Exception as gps_exc:
                        exif["GPS_Error"] = str(gps_exc)

                    # --- All other tags
                    for tag_id, val in raw.items():
                        if tag_id == GPS_IFD_TAG:
                            continue  # already handled above
                        tag = ExifTags.TAGS.get(tag_id, str(tag_id))
                        if val is None:
                            continue
                        elif isinstance(val, bytes):
                            decoded = val.decode("utf-8", errors="ignore").strip()
                            if decoded:
                                exif[str(tag)] = decoded
                        else:
                            exif[str(tag)] = str(val)
            except Exception as exif_exc:
                result["exif_error"] = str(exif_exc)

            # Elevate device fingerprint to top-level fields
            if "Make" in exif:
                result["device_make"] = exif["Make"]
            if "Model" in exif:
                result["device_model"] = exif["Model"]
            if "Software" in exif:
                result["software"] = exif["Software"]
            if "DateTimeOriginal" in exif:
                result["date_taken"] = exif["DateTimeOriginal"]

            # Reverse geocode GPS if present
            if gps_lat is not None and gps_lon is not None:
                location = _reverse_geocode(gps_lat, gps_lon)
                if location:
                    exif["GPS_Location"] = location
                    result["location"] = location

            result["metadata"] = exif
    except Exception as exc:
        result["error"] = str(exc)
    return result


def _audio(path: Path) -> dict:
    result = {"file_type": "audio", **_file_base(path)}
    try:
        import mutagen
        audio = mutagen.File(path, easy=True)
        if audio is None:
            result["error"] = "Could not parse audio file"
            return result
        info: dict = {}
        if hasattr(audio, "info"):
            ai = audio.info
            if hasattr(ai, "length"):
                info["duration_seconds"] = round(float(ai.length), 2)
            if hasattr(ai, "bitrate"):
                info["bitrate_kbps"] = getattr(ai, "bitrate", None)
            if hasattr(ai, "sample_rate"):
                info["sample_rate_hz"] = getattr(ai, "sample_rate", None)
            if hasattr(ai, "channels"):
                info["channels"] = getattr(ai, "channels", None)
        tags = {}
        for k, v in audio.items():
            tags[k] = str(v[0]) if isinstance(v, list) and v else str(v)
        result["audio_info"] = info
        result["metadata"] = tags
    except ImportError:
        result["error"] = "mutagen not installed. Run: pip install mutagen"
    except Exception as exc:
        result["error"] = str(exc)
    return result


def _video(path: Path) -> dict:
    result = {"file_type": "video", **_file_base(path)}
    try:
        from hachoir.parser import createParser  # type: ignore
        from hachoir.metadata import extractMetadata  # type: ignore
        parser = createParser(str(path))
        if parser:
            with parser:
                metadata = extractMetadata(parser)
            if metadata:
                items = {}
                for line in metadata.exportPlaintext():
                    if ": " in line:
                        k, _, v = line.partition(": ")
                        items[k.strip("- ").strip()] = v.strip()
                result["metadata"] = items
                return result
    except ImportError:
        pass
    except Exception:
        pass
    result["metadata"] = {}
    result["note"] = "Install hachoir for deeper video metadata: pip install hachoir"
    return result


def _pdf(path: Path) -> dict:
    result = {"file_type": "pdf", **_file_base(path)}
    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(str(path))
        meta = reader.metadata or {}
        result["page_count"] = len(reader.pages)
        result["encrypted"] = reader.is_encrypted
        result["metadata"] = {k.lstrip("/"): str(v) for k, v in meta.items()}
    except ImportError:
        result["error"] = "pypdf not installed. Run: pip install pypdf"
    except Exception as exc:
        result["error"] = str(exc)
    return result


def _docx(path: Path) -> dict:
    result = {"file_type": "docx", **_file_base(path)}
    try:
        from docx import Document  # type: ignore
        doc = Document(str(path))
        cp = doc.core_properties
        result["metadata"] = {
            "author": cp.author or "",
            "last_modified_by": cp.last_modified_by or "",
            "created": str(cp.created) if cp.created else "",
            "modified": str(cp.modified) if cp.modified else "",
            "title": cp.title or "",
            "subject": cp.subject or "",
            "description": cp.description or "",
            "revision": cp.revision,
            "word_count": sum(len(p.text.split()) for p in doc.paragraphs),
        }
    except ImportError:
        result["error"] = "python-docx not installed. Run: pip install python-docx"
    except Exception as exc:
        result["error"] = str(exc)
    return result


def _xlsx(path: Path) -> dict:
    result = {"file_type": "xlsx", **_file_base(path)}
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        cp = wb.properties
        result["sheet_count"] = len(wb.sheetnames)
        result["sheets"] = wb.sheetnames
        result["metadata"] = {
            "creator": cp.creator or "",
            "last_modified_by": cp.lastModifiedBy or "",
            "created": str(cp.created) if cp.created else "",
            "modified": str(cp.modified) if cp.modified else "",
            "title": cp.title or "",
            "subject": cp.subject or "",
        }
    except ImportError:
        result["error"] = "openpyxl not installed. Run: pip install openpyxl"
    except Exception as exc:
        result["error"] = str(exc)
    return result


_DISPATCH = {
    ".jpg": _image, ".jpeg": _image, ".png": _image, ".gif": _image,
    ".bmp": _image, ".tiff": _image, ".tif": _image, ".webp": _image,
    ".mp3": _audio, ".flac": _audio, ".wav": _audio, ".ogg": _audio,
    ".aac": _audio, ".m4a": _audio, ".wma": _audio,
    ".mp4": _video, ".avi": _video, ".mov": _video, ".mkv": _video,
    ".wmv": _video, ".flv": _video, ".webm": _video,
    ".pdf": _pdf,
    ".docx": _docx,
    ".xlsx": _xlsx, ".xls": _xlsx,
}


def analyse_file(path: Path) -> dict:
    """Dispatch to the appropriate extractor based on file extension."""
    ext = path.suffix.lower()
    handler = _DISPATCH.get(ext)
    if handler:
        return handler(path)
    result = {"file_type": "unknown", **_file_base(path)}
    result["note"] = f"No metadata extractor available for {ext or 'unknown'} files"
    return result


SUPPORTED_EXTENSIONS = set(_DISPATCH.keys())
